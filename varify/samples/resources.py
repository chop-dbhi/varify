import functools
import openpyxl
import logging
import requests
import itertools
from django.core import management
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.conf.urls import patterns, url
from django.core.cache import cache
from django.http import HttpResponse, Http404
from django.core.exceptions import PermissionDenied
from django.core.urlresolvers import reverse
from django.views.decorators.cache import never_cache
from django.conf import settings
from guardian.shortcuts import get_objects_for_user
from openpyxl.shared.exc import OpenModeError, InvalidFileException
from preserialize.serialize import serialize
from restlib2 import resources
from restlib2.http import codes
from serrano.resources.base import ThrottledResource
from varify.variants.resources import VariantResource
from varify import api
from vdw.assessments.models import Assessment
from vdw.genome.models import Chromosome
from vdw.samples.models import Sample, Result, ResultScore, ResultSet
from vdw.variants.models import Variant
from .forms import ResultSetForm

log = logging.getLogger(__name__)
OPENPYXL_MAJOR_VERSION = int(openpyxl.__version__[0])


def get_cell_value(cell):
    """
    Convenience method for getting the value of an openpyxl cell

    This is necessary since the value property changed from internal_value
    to value between version 1.* and 2.*.
    """
    if OPENPYXL_MAJOR_VERSION > 1:
        return cell.value

    return cell.internal_value


def set_cell_value(cell, value):
    """
    Convenience method for setting the value of an openpyxl cell

    This is necessary since the value property changed from internal_value
    to value between version 1.* and 2.*.
    """
    if OPENPYXL_MAJOR_VERSION > 1:
        cell.value = value
    else:
        cell.internal_value = value


def sample_posthook(instance, data, request):
    uri = request.build_absolute_uri

    data['_links'] = {
        'self': {
            'rel': 'self',
            'href': uri(reverse('api:samples:sample',
                                args=[instance.pk])),
        },
        'variants': {
            'rel': 'related',
            'href': uri(reverse('api:samples:variants',
                                args=[instance.pk])),
        },
        'variant-sets': {
            'rel': 'related',
            'href': uri(reverse('api:samples:variant-sets',
                                args=[instance.pk])),
        }
    }

    return data


class SampleBaseResource(ThrottledResource):
    model = Sample

    template = api.templates.Sample

    def get_queryset(self, request, **kwargs):
        projects = get_objects_for_user(request.user, 'samples.view_project')
        return self.model.objects.select_related('batch', 'project')\
            .filter(project__in=projects)

    def get_object(self, request, pk):
        if not hasattr(request, 'instance'):
            queryset = self.get_queryset(request)

            try:
                instance = queryset.get(pk=pk)
            except self.model.DoesNotExist:
                instance = None

            request.instance = instance

        return request.instance


class SamplesResource(SampleBaseResource):
    def get(self, request):
        samples = self.get_queryset(request)
        posthook = functools.partial(sample_posthook, request=request)
        return serialize(samples, posthook=posthook, **self.template)


class SampleResource(SampleBaseResource):
    def is_not_found(self, request, response, pk):
        return not self.get_object(request, pk=pk)

    @api.cache_resource
    def get(self, request, pk):
        instance = self.get_object(request, pk=pk)
        posthook = functools.partial(sample_posthook, request=request)
        return serialize(instance, posthook=posthook, **self.template)


class NamedSampleResource(ThrottledResource):
    "Resource for looking up a sample by project, batch, and sample name"
    model = Sample

    template = api.templates.Sample

    # Bypass authorization check imposed by Serrano's AUTH_REQUIRED setting
    def __call__(self, *args, **kwargs):
        return resources.Resource.__call__(self, *args, **kwargs)

    def is_not_found(self, request, response, project, batch, sample):
        try:
            instance = self.model.objects.get(project__name=project,
                                              batch__name=batch, name=sample)
        except self.model.DoesNotExist:
            return True

        request.instance = instance
        return False

    def get(self, request, project, batch, sample):
        posthook = functools.partial(sample_posthook, request=request)
        return serialize(request.instance, posthook=posthook, **self.template)


class SampleResultBaseResource(ThrottledResource):
    def is_not_found(self, request, response, pk):
        projects = get_objects_for_user(request.user, 'samples.view_project')
        return not Sample.objects.filter(project__in=projects, pk=pk).exists()


class SampleResultsResource(SampleResultBaseResource):
    "Paginated view of results for a sample."
    model = Result

    template = api.templates.SampleResult

    def get(self, request, pk):
        uri = request.build_absolute_uri
        page = request.GET.get('page', 1)

        related = ['sample', 'variant', 'variant__chr', 'genotype']
        results = self.model.objects.select_related(*related)\
            .filter(sample__pk=pk)

        # Paginate the results
        paginator = Paginator(results, api.PAGE_SIZE)

        try:
            page = page = paginator.page(page)
        except PageNotAnInteger:
            page = paginator.page(1)
        except EmptyPage:
            page = paginator.page(paginator.num_pages)

        resp = {
            'result_count': paginator.count,
            'results': serialize(page.object_list, **self.template),
        }

        # Augment the links
        for obj in resp['results']:
            obj['_links'] = {
                'self': {
                    'rel': 'self',
                    'href': uri(reverse('api:samples:variant',
                                        kwargs={'pk': obj['id']}))
                },
                'sample': {
                    'rel': 'related',
                    'href': uri(reverse('api:samples:sample',
                                        kwargs={'pk': obj['sample']['id']}))
                },
                'variant': {
                    'rel': 'related',
                    'href': uri(reverse('api:variants:variant',
                                        kwargs={'pk': obj['variant_id']})),
                }
            }
            obj.pop('variant_id')

        links = {
            'self': {
                'rel': 'self',
                'href': uri(reverse('api:samples:variants',
                                    kwargs={'pk': pk})),
            }
        }

        if page.number != 1:
            links['prev'] = {
                'rel': 'prev',
                'href': uri('{0}?page={1}'.format(
                    reverse('api:samples:variants', kwargs={'pk': pk}),
                    str(page.number - 1)))
            }

        if page.number < paginator.num_pages - 1:
            links['next'] = {
                'rel': 'next',
                'href': uri('{0}?page={1}'.format(
                    reverse('api:samples:variants', kwargs={'pk': pk}),
                    str(page.number + 1)))
            }

        if links:
            resp['_links'] = links

        return resp


class SampleResultResource(ThrottledResource):
    model = Result

    template = api.templates.SampleResultVariant

    def is_not_found(self, request, response, pk):
        return not self.model.objects.filter(pk=pk).exists()

    def _cache_data(self, request, pk, key):
        uri = request.build_absolute_uri
        related = ['sample', 'variant', 'genotype', 'score']

        try:
            result = self.model.objects.select_related(*related).get(pk=pk)
        except self.model.DoesNotExist:
            raise Http404

        data = serialize(result, **self.template)

        data['_links'] = {
            'self': {
                'rel': 'self',
                'href': uri(reverse('api:samples:variant',
                                    kwargs={'pk': data['id']}))
            },
            'sample': {
                'rel': 'related',
                'href': uri(reverse('api:samples:sample',
                                    kwargs={'pk': data['sample']['id']}))
            },
            'variant': {
                'rel': 'related',
                'href': uri(reverse('api:variants:variant',
                                    kwargs={'pk': data['variant_id']})),
            }
        }

        # Integrate the Variant resource data
        data['variant'] = VariantResource.get(request, data['variant_id'])
        data.pop('variant_id')

        try:
            score = ResultScore.objects.get(result=result)
            data['score'] = {
                'score': score.score,
                'rank': score.rank,
            }
        except ResultScore.DoesNotExist:
            pass

        cache.set(key, data, timeout=api.CACHE_TIMEOUT)
        return data

    def get(self, request, pk):
        key = api.cache_key(self.model, pk)
        data = cache.get(key)

        if data is None:
            data = self._cache_data(request, pk, key)

        try:
            assessment = Assessment.objects.get(sample_result=pk,
                                                user=request.user.id)
            data['assessment'] = serialize(assessment,
                                           **api.templates.ResultAssessment)
        except Assessment.DoesNotExist:
            data['assessment'] = {}

        return data

    post = get


class ResultsResource(ThrottledResource):
    template = api.templates.SampleResultVariant

    def post(self, request):
        if (not request.data.get('ids') or
                not isinstance(request.data['ids'], list)):
            return HttpResponse(status=codes.unprocessable_entity,
                                content='Array of "ids" is required')

        data = []
        resource = SampleResultResource()
        for id in request.data['ids']:
            data.append(resource.get(request, id))

        return data


class PhenotypeResource(ThrottledResource):
    def get(self, request, sample_id):
        recalculate = request.GET.get('recalculate_rankings')

        if recalculate == 'true':
            try:
                management.call_command('samples', 'gene-ranks', sample_id,
                                        force=True)
            except Exception:
                log.exception('Error recalculating gene rankings')
                return self.render(request, {
                    'message': 'Error recalculating gene rankings',
                }, status=codes.server_error)

        endpoint = getattr(settings, 'PHENOTYPE_ENDPOINT', None)

        if not endpoint:
            log.error('PHENOTYPE_ENDPOINT setting could not be found.')
            return self.render(request, '', status=codes.server_error)

        endpoint = endpoint.format(sample_id)

        try:
            response = requests.get(endpoint, cert=(settings.VARIFY_CERT,
                                    settings.VARIFY_KEY), verify=False)
        except requests.exceptions.SSLError:
            raise PermissionDenied
        except requests.exceptions.ConnectionError:
            return self.render(request, '', status=codes.server_error)
        except requests.exceptions.RequestException:
            raise Http404

        # If anything at all goes wrong in the sample lookup or json parsing
        # then just abandon all hope and return the content from the orignal
        # response.
        try:
            sample = Sample.objects.get(label=sample_id)
            data = response.json()
            data['phenotype_modified'] = sample.phenotype_modified
            return data
        except Exception:
            pass

        return response.content


class PedigreeResource(ThrottledResource):
    def get(self, request, year, month, day, name):
        endpoint = getattr(settings, 'PEDIGREE_ENDPOINT', None)

        if not endpoint:
            log.error('PEDIGREE_ENDPOINT setting could not be found.')
            return self.render(request, '', status=codes.server_error)

        endpoint = endpoint.format(year, month, day, name)

        try:
            pedigree_response = requests.get(
                endpoint, cert=(settings.VARIFY_CERT, settings.VARIFY_KEY),
                verify=False)
        except requests.exceptions.SSLError:
            raise PermissionDenied
        except requests.exceptions.ConnectionError:
            return self.render(request, '', status=codes.server_error)
        except requests.exceptions.RequestException:
            raise Http404

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = \
            'attachment; filename="{0}-{1}-{2}-{3}"'.format(
                year, month, day, name)

        response.write(pedigree_response.content)

        return response


class SampleResultSetsResource(ThrottledResource):
    model = ResultSet

    template = api.templates.SimpleResultSet

    INSERTION = 'Insertion'
    DELETION = 'Deletion'

    supported_content_types = (
        'application/json',
        'multipart/form-data',
    )

    def get_change_type(self, ref, a1, a2):
        """
        Given ref, allele1, and allele2, returns the type of change.
        The only case of an amino acid insertion is when the ref is
        represented as a '.'.
        """
        if ref == '.':
            return self.INSERTION
        elif a1 == '.' or a2 == '.':
            return self.DELETION

    def get_queryset(self, request, pk):
        return self.model.objects.filter(sample__pk=pk)

    def get(self, request, pk):
        queryset = self.get_queryset(request, pk)
        return serialize(queryset, **self.template)

    # Consume the variants excel file and search the database for
    # matching queries.
    def _create_from_file(self, request, pk, instance):
        sample = instance.sample

        # Try opening the excel file that was uploaded.
        try:
            variant_book = request.FILES['source']
            workbook = openpyxl.load_workbook(variant_book,
                                              use_iterators=True)
        except (OpenModeError, InvalidFileException):
            log.exception('Error opening workbook when creating result set')
            return self.render(request, 'Could not process the file. Make '
                               'sure that the file is a valid excel file',
                               status=codes.unprocessable_entity)

        sheet = workbook.get_active_sheet()

        start_row = 0
        fields = []

        # Skip the header information of the data file (if there is any)
        # and find the row where the data starts by finding the first instance
        # of a column title.
        for row in sheet.iter_rows():
            if get_cell_value(row[0]) == 'Chromosome':
                # Retrieve all the column titles from the sheet.
                fields = [get_cell_value(r).lower() for r in row]
                break
            start_row += 1

        allele1_index = fields.index('allele 1') \
            if 'allele 1' in fields else None
        allele2_index = fields.index('allele 2') \
            if 'allele 2' in fields else None
        dbsnp_index = fields.index('dbsnp')
        ref_index = fields.index('reference')
        start_index = fields.index('start')
        chr_index = fields.index('chromosome')

        matches = set()
        no_matches = []

        # Verison 1.6.1 of openpyxl does not support slicing thus, a workaround
        # proposed here:
        #
        #       https://bitbucket.org/ericgazoni/openpyxl/issue/273/allow-slicing-in-iterableworksheet      # noqa
        # TODO: Speed this up. This code is painfully slow, takes upwards of
        # 4 minutes to process ~5,000 records.
        for row in itertools.islice(sheet.iter_rows(), start_row + 1, None):
            # Retrieve rsid, ref, start and chr from our spreadsheet
            # so we can retrieve objects from our database.
            dbsnp_value = get_cell_value(row[dbsnp_index])
            ref_value = get_cell_value(row[ref_index])
            start_value = int(get_cell_value(row[start_index]))

            # In the case that the position changes and there was no matching
            # query. Save the original to return back to the user.
            old_position = start_value

            chr_label = get_cell_value(row[chr_index])

            chr_match = Chromosome.objects.get(value=chr_label)

            allele1 = get_cell_value(row[allele1_index]) \
                if allele1_index else None
            allele2 = get_cell_value(row[allele2_index]) \
                if allele1_index else None

            is_found = False

            # Fetch the variant w/ matching dbsnp id and the
            # result which have the variant.
            if dbsnp_value:
                try:
                    result = Result.objects.get(sample=sample,
                                                variant__rsid=dbsnp_value)
                    matches.add(result)
                    is_found = True
                except Result.DoesNotExist:
                    pass

            if allele1 and not is_found:
                # Determine the type of change. This is required to resolve
                # the case where two variants have the same genomic start
                # position.
                change_type = self.get_change_type(ref_value, allele1, allele2)

                if change_type:
                    # In the case of an insertion or a deletion, the start
                    # position is always decremented by 1 and the query is
                    # is made using only the position and chromosome.
                    start_value -= 1

                else:
                    # Now using the two alleles, construct the alt.
                    if allele1 == ref_value:
                        alt_value = allele2
                    elif allele1 != allele2:
                        alt_value = '{0},{1}'.format(allele1, allele2)
                    else:
                        alt_value = allele1

                # Retrieve variant needed for our result query.
                try:
                    correct_variant = None

                    # If there was a deletion or an insertion, make the query
                    # using the position and the chromosome and resolve any
                    # conflicts.
                    if change_type:
                        variant_match = Variant.objects.filter(
                            pos=start_value,
                            chr=chr_match)

                        if len(variant_match) == 1:
                            correct_variant = variant_match[0]

                        elif len(variant_match) > 1:
                            for match in variant_match:
                                # In the case of an insertion, the variants
                                # reference string should have increased by 1
                                # amino acid.
                                if change_type == self.INSERTION and \
                                        len(match.ref) == len(ref_value) + 1:
                                    correct_variant = match
                                    break
                                # In the case of a deletion (where the value
                                # of the reference is '.'. The reference should
                                # have a length of 1 amino acid.
                                elif change_type == self.DELETION and \
                                        len(match.ref) == 1:
                                    correct_variant = match
                                    break

                    else:
                        correct_variant = Variant.objects.get(
                            pos=start_value,
                            ref=ref_value,
                            alt=alt_value,
                            chr=chr_match)

                    result = Result.objects.get(
                        sample=sample,
                        variant=correct_variant)

                    matches.add(result)
                    is_found = True
                except (Variant.DoesNotExist, Result.DoesNotExist):
                    pass

            # Queries were unsuccessful so save this row to return back to the
            # user to edit/fix them.
            if not is_found:
                no_matches.append({
                    'chr': chr_label,
                    'start': old_position,
                    'ref': ref_value,
                    'allele1': allele1,
                    'allele2': allele2,
                    'dbsnp': dbsnp_value,
                })

        instance.bulk(matches)

        content = serialize(instance, **self.template)
        content['num_total_records'] = len(matches) + len(no_matches)
        content['invalid_records'] = no_matches

        return self.render(request, content=content, status=codes.created)

    def _create_from_context(self, request, instance):
        context = self.get_context(request)
        results = context.apply()
        instance.bulk(results)

        content = serialize(instance, **self.template)
        return self.render(request, content=content, status=codes.created)

    def post(self, request, pk):
        if 'source' in request.FILES and request.FILES['source']:
            data = request.POST
        else:
            data = request.data

        data['sample'] = Sample.objects.get(pk=pk).id

        form = ResultSetForm(data)

        if form.is_valid():
            instance = form.save(commit=False)
            instance.user = request.user
            instance.save()

            # Determine which type of post request we are dealing with and then
            # pass the necessary data on to the internal create methods. We
            # either create a variant set from an uploaded file or we create
            # one from the request user's context.
            if 'source' in request.FILES and request.FILES['source']:
                return self._create_from_file(request, pk, instance)
            else:
                return self._create_from_context(request, instance)
        else:
            log.error('Invalid variant set input. Form errors: {0}'.format(
                dict(form.errors)))
            return self.render(request, {'message': 'Request data is invalid'},
                               codes.unprocessable_entity)


class SampleResultSetResource(SampleResultSetsResource):
    model = ResultSet

    template = api.templates.ResultSet

    def get_object(self, request, pk):
        if not hasattr(request, 'instance'):
            try:
                instance = self.model.objects.get(pk=pk)
            except self.model.DoesNotExist:
                instance = None

            request.instance = instance

        return request.instance

    def is_not_found(self, request, response, pk):
        return not self.get_object(request, pk)

    def get(self, request, pk):
        instance = self.get_object(request, pk)
        data = serialize(instance, **self.template)

        for i in range(len(data['results'])):
            data['results'][i]['variant'] = VariantResource.get(
                request, data['results'][i]['variant_id'])
            data['results'][i].pop('variant_id')

            result_id = data['results'][i]['id']
            data['results'][i]['num_assessments'] = Assessment.objects.filter(
                sample_result__id=result_id,
                sample_result__resultset__id=pk).count()

            try:
                assessment = Assessment.objects.get(
                    sample_result__id=result_id,
                    sample_result__resultset__id=pk,
                    user=request.user.id)
                data['results'][i]['assessment'] = \
                    serialize(assessment, **api.templates.ResultAssessment)
            except Assessment.DoesNotExist:
                data['results'][i]['assessment'] = {}

        return data


sample_resource = never_cache(SampleResource())
samples_resource = never_cache(SamplesResource())
named_sample_resource = never_cache(NamedSampleResource())

sample_results_resource = never_cache(SampleResultsResource())
sample_result_resource = never_cache(SampleResultResource())
results_resource = never_cache(ResultsResource())

sample_result_sets_resource = never_cache(SampleResultSetsResource())
sample_result_set_resource = never_cache(SampleResultSetResource())

phenotype_resource = never_cache(PhenotypeResource())
pedigree_resource = never_cache(PedigreeResource())

urlpatterns = patterns(
    '',

    url(r'^$',
        samples_resource,
        name='samples'),

    url(r'^(?P<pk>\d+)/$',
        sample_resource,
        name='sample'),

    url(r'^(?P<pk>\d+)/variants/$',
        sample_results_resource,
        name='variants'),

    url(r'^variants/(?P<pk>\d+)/$',
        sample_result_resource,
        name='variant'),

    url(r'^variants/$',
        results_resource,
        name='results_resource'),

    url(r'^(?P<pk>\d+)/variants/sets/$',
        sample_result_sets_resource,
        name='variant-sets'),

    url(r'^variants/sets/(?P<pk>\d+)/$',
        sample_result_set_resource,
        name='variant-set'),

    url(r'^(?P<sample_id>.+)/phenotypes/$',
        phenotype_resource,
        name='phenotype'),

    url(r'^pedigrees/(?P<year>\d+)/(?P<month>\d+)/(?P<day>\d+)/(?P<name>.+)$',
        pedigree_resource,
        name='pedigree'),

    # At the bottom to prevent matching on such variable parameters
    url(r'^(?P<project>.+)/(?P<batch>.+)/(?P<sample>.+)/$',
        named_sample_resource,
        name='named_sample'),
)
