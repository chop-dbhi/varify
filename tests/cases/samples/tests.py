import json
import os
from openpyxl import Workbook, cell
from restlib2.http import codes
from vdw.samples.models import Project
from guardian.shortcuts import assign
from ..base import AuthenticatedBaseTestCase


class VariantUploadResourceTestCase(AuthenticatedBaseTestCase):
    fixtures = ['initial_variants.json']

    def setUp(self):
        super(VariantUploadResourceTestCase, self).setUp()

    def test_post(self):
        book = Workbook()
        sheet1 = book.get_active_sheet()
        sheet1.title = 'Variants List'
        fields = ['Chromosome', 'Start', 'Reference', 'Allele 1', 'Allele 2',
                  'dbSNP']

        # Create variants to cover all edge cases, including the case
        # where there is a variation at the same genomic position.
        # Also consider the case where the dbSNP id is incorrect.
        variants = [['1', '20000', '.', 'AC', 'AC', ''],
                    ['1', '20000', 'A', 'A', '.', ''],
                    ['3', '20002', 'GAT', '.', '.', 'rs9160301'],
                    ['1', '20003', '.', '.', 'TTTCTT', ''],
                    ['3', '20004', 'A', 'C', 'C', 'rs916000'],
                    ['1', '20007', 'GTCATTGGAACAGTC', '.',
                     'GTCATTGGAACAGTC', '']]

        # Write our fields in first.
        for i in range(0, 6):
            sheet1.cell(row=0, column=i).set_value_explicit(
                value=fields[i], data_type=cell.Cell.TYPE_STRING)

        # Write the variants to the excel sheet.
        row = 1
        for v in variants:
            for col in range(0, 6):
                sheet1.cell(row=row, column=col).set_value_explicit(
                    value=v[col], data_type=cell.Cell.TYPE_STRING)
            row += 1

        book.save('variantList.xlsx')

        with open('variantList.xlsx') as fp:
            # Assign permissions for test user to access the project.
            p = Project.objects.get(pk=1)
            assign('samples.view_project', self.user, p)

            response = self.client.post('/api/samples/9/variants/sets/',
                                        {'name': 'variants',
                                         'source': fp})
            response_obj = json.loads(response.content)

            # An array of matching variants are returned. Make sure
            # that exactly 5 were found and 1 was added to the
            # list of unmatched variants.
            self.assertEqual(response.status_code, codes.created)
            self.assertEqual(response_obj['num_total_records'], 6)
            self.assertEqual(response_obj['count'], 5)
            self.assertEqual(len(response_obj['invalid_records']), 1)

        os.remove('variantList.xlsx')
